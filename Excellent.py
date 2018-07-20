from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.comments import Comment
from openpyxl.utils import coordinate_from_string, column_index_from_string
from math import exp, expm1

import os  
import ntpath
import ChartFinder

reserved_for_chart = exp(exp(exp(1)))
yellow = "FFFFFF99"
red = "FFFF3333"
orange = "FFFF6600"
pink = "FFFF00CC"
purple = "FF9900FF"
blue = "FF6666FF"
green = "FF99FF66"
magenta = "FFFF33FF"
black = "FF000000"
white = "FFFFFFFF"
nofill = "00000000"
grey5 = "FF999999"

tolrnc_percent = 0.0001

def is_around(target,value,tolrnc_percent):
	if (type(value) == float or type(value) == int or type(value) == long ) and (type(target) == float or type(target) == int or type(target) == long ):
		value = float(value)
		target = float(target)
		if abs((value-target)/target*100.0) < tolrnc_percent:
			return True
		else:
			return False
	elif type(value) == unicode and type(target) == unicode:
		if value == target:
			return True
		else:
			return False

def is_cell_font_colored( FileName,sheet,cell,shade ):
	wb = load_workbook(filename = FileName, read_only=True)
	ws = wb.get_sheet_by_name(sheet)
	if ws[cell].font.color.rgb == shade:
		return True
		print "Red Spotted"
	else:
		return False
		print "Non-Red Spotted"

def how_many_numeric_genuine_match_in_sheet( FileName,sheet,target, ):
	wb = load_workbook(filename = FileName, data_only=True,read_only=True)
	wb_formula = load_workbook(filename = FileName, data_only=False,read_only=True)
	i=0
	ws = wb.get_sheet_by_name(sheet)
	ws_formula = wb_formula.get_sheet_by_name(sheet)
	for row in ws.iter_rows():
		for cell in row:
			if type(cell.value) == float or type(cell.value) == int or type(cell.value) == long:
				if is_around(target,float(cell.value),tolrnc_percent):
					i=i+1
					#letrow = letrow_from_colrow_nn(cell.column,cell.row)
					#if is_cell_font_colored( FileName,sheet,letrow,red ) == True:
					#	if(ws_formula[letrow].value != ws[letrow].value):
					#		i=i+1
					#elif is_cell_font_colored( FileName,sheet,letrow,red ) == False:

	return i

def how_many_match_in_sheet( FileName,sheet,target,stop_string_array ):

	wb = load_workbook(filename = FileName, data_only=True,read_only=True)
	ws = wb.get_sheet_by_name(sheet)

	if len(stop_string_array) == 0:
		i=0
		for row in ws.iter_rows():
			for cell in row:
				if is_around(target,cell.value,tolrnc_percent):
					i=i+1
	else:
		did_inner_loop_break = False
		i=0
		for row in ws.iter_rows():
			for cell in row:
				for iter in range(0,len(stop_string_array)):
					if is_around(stop_string_array[iter],cell.value,tolrnc_percent):
						#print stop_string_array[iter]
						did_inner_loop_break = True
						break
				if is_around(target,cell.value,tolrnc_percent):
					i=i+1
					#print "Found:", cell.value," Target: ",target
				if did_inner_loop_break == True:
					break
			if did_inner_loop_break == True:
				break
	
	
	return i

def how_many_none_numeric_match_in_sheet( FileName,sheet,value ):
	wb = load_workbook(filename = FileName,read_only=True)
	i=0
	ws = wb.get_sheet_by_name(sheet)
	for row in ws.iter_rows():
		for cell in row:
			if type(cell.value) != float and type(cell.value) != int and type(cell.value) != long:
				if value == cell.value:
					i=i+1
	return i

def get_cell_value( FileName,sheet,cell ):
	wb = load_workbook(filename = FileName, data_only=True, read_only=True)
	ws = wb.get_sheet_by_name(sheet)
		
	return ws[cell].value

def get_cell_formula( FileName,sheet,cell ):
	wb = load_workbook(filename = FileName, data_only=False, read_only=True)
	ws = wb.get_sheet_by_name(sheet)
		
	return ws[cell].value

def is_cell_highlighted( FileName,sheet,cell,shade ):
	wb = load_workbook(filename = FileName, read_only=True)
	ws = wb.get_sheet_by_name(sheet)
	if ws[cell].fill.start_color.index == shade:
		return True
	else:
		return False	

def find_highlighted_cells( FileName, shade ):
	wb = load_workbook(filename = FileName, read_only=True)
	for ws in wb.worksheets:
		for row in ws.iter_rows():
			for cell in row:
				if cell.value:
					if cell.fill.start_color.index == shade:
						print ws,cell.row,cell.column,"(",get_column_letter(cell.column),cell.row,")"
	return

def find_highlighted_cells_in_sheet( FileName,sheet, shade ):
	wb = load_workbook(filename = FileName, read_only=True)
	ws = wb.get_sheet_by_name(sheet)
	for row in ws.iter_rows():
		for cell in row:
			if cell.value:
				if cell.fill.start_color.index == shade:
					print ws,cell.row,cell.column,"(",get_column_letter(cell.column),cell.row,")"
	return

def how_many_highlighted_cells_in_book( FileName, shade ):
	wb = load_workbook(filename = FileName, read_only=True)
	i=0
	for ws in wb.worksheets:
		for row in ws.iter_rows():
			for cell in row:
				if cell.value:
					if cell.fill.start_color.index == shade:
						i = i + 1
	return i

def how_many_highlighted_cells_in_sheet( FileName,sheet, shade ):
	wb = load_workbook(filename = FileName, read_only=True)
	ws = wb.get_sheet_by_name(sheet)
	i=0
	for row in ws.iter_rows():
		for cell in row:
			if cell.value:
				if cell.fill.start_color.index == shade:
					i = i + 1
	return i

def get_highlighted_cell_values_in_sheet( FileName,sheet, shade ):
	wb = load_workbook(filename = FileName, data_only= True,read_only=True)
	ws = wb.get_sheet_by_name(sheet)
	highlighted=[]
	for row in ws.iter_rows():
		for cell in row:
			if cell.value:
				if cell.fill.start_color.index == shade:
					highlighted.append(cell.value)
	return highlighted

def get_highlighted_cell_letrows_in_sheet( FileName,sheet, shade ):
	wb = load_workbook(filename = FileName, data_only= True,read_only=True)
	ws = wb.get_sheet_by_name(sheet)
	highlighted=[]
	for row in ws.iter_rows():
		for cell in row:
			if cell.value:
				if cell.fill.start_color.index == shade:
					highlighted.append(str(get_column_letter(cell.column))+str(cell.row))
	return highlighted

def get_highlighted_cell__in_sheet( FileName,sheet, shade ):
	wb = load_workbook(filename = FileName, data_only= True,read_only=True)
	ws = wb.get_sheet_by_name(sheet)
	highlighted=[]
	for row in ws.iter_rows():
		for cell in row:
			if cell.value:
				if cell.fill.start_color.index == shade:
					highlighted.append(cell.value)
	return highlighted

def get_highlighted_cell_cols_in_sheet( FileName,sheet, shade ):
	wb = load_workbook(filename = FileName, data_only= True,read_only=True)
	ws = wb.get_sheet_by_name(sheet)
	highlighted=[]
	for row in ws.iter_rows():
		for cell in row:
			if cell.value:
				if cell.fill.start_color.index == shade:
					highlighted.append(cell.column)
	return highlighted

def get_highlighted_cell_rows_in_sheet( FileName,sheet, shade ):
	wb = load_workbook(filename = FileName, data_only= True,read_only=True)
	ws = wb.get_sheet_by_name(sheet)
	highlighted=[]
	for row in ws.iter_rows():
		for cell in row:
			if cell.value:
				if cell.fill.start_color.index == shade:
					highlighted.append(cell.row)
	return highlighted

def how_many_this_font_coor_in_sheet( FileName,sheet, shade ):
	wb = load_workbook(filename = FileName, read_only=True)
	ws = wb.get_sheet_by_name(sheet)
	for row in ws.iter_rows():
		for cell in row:
			if cell.value:
				if cell.font.color.rgb == red:
					print cell.row,cell.column,"(",get_column_letter(cell.column),cell.row,")"
	return

def find_red_font( FileName ):
	wb = load_workbook(filename = FileName, read_only=True)
	for ws in wb.worksheets:
		for row in ws.iter_rows():
			for cell in row:
				if cell.value:
					if cell.font.color.rgb == red:
						print cell.row,cell.column,"(",get_column_letter(cell.column),cell.row,")"
	return

def get_comment(Filename,sheet,cell):
	book = load_workbook(filename = Filename)
	sheet=book.get_sheet_by_name(sheet)
	#print sheet[cell].comment.text
	return sheet[cell].comment.text

def get_sheet_names_b(Filename):
	book = load_workbook(filename = Filename, read_only=True)
	#print book.get_sheet_names()
	return book.get_sheet_names()	#outputs a list

def get_nth_sheet_bn(Filename,n):
	book = load_workbook(filename = Filename, read_only=True)
	sheets_array=book.get_sheet_names()
	#print book.get_sheet_names()
	return sheets_array[n-1]

def get_sheet_number(Filename,sheet):
	book = load_workbook(filename = Filename, read_only=True)
	sheets_array=book.get_sheet_names()
	for i in range(0,len(sheets_array)):
		if sheet == sheets_array[i]:
			return i+1

def count_sheets_b( FileName ):
	wb = load_workbook(filename = FileName, read_only=True)
	i = 0
	for ws in wb.worksheets:
		i = i + 1
	return i

def letrow_from_colrow_nn(col,row):
	#print get_column_letter(col)+str(row)
	ans = str(get_column_letter(col))+str(row)
	return ans

def colrow_from_letrow_s(string):
	letrow = coordinate_from_string(string) 
	return column_index_from_string(letrow[0]),letrow[1]	# for B4 returns (2,4)

def sum_array(array):
	sum = 0.0
	for i in range(0,len(array)):
		sum = sum + array[i]
	return sum

def get_students_first_last(string):
	name = []
	splitted_string = string.split()
	name.append(splitted_string[0])
	lastName = splitted_string[1]
	name.append(lastName[:-1])
	return name

#==================================================================================================#

def list_sheets_all_books(submission_folder):
	i = 0
	j = 0
	for subdir, dirs, files in os.walk(submission_folder):
	    for file in sorted(files):
		ext = os.path.splitext(file)[-1].lower()
		if ext == ".xlsx":
			i =  i + 1
			print i, len(get_sheet_names_b(file)),file
	print 
	for file in sorted(files):
		ext = os.path.splitext(file)[-1].lower()
		if ext == ".xlsx":
			j =  j + 1
			print j, get_sheet_names_b(file)
	return


def analyze_reference_sheet(reference_book, reference_sheet):
	hilites = get_highlighted_cell_values_in_sheet( reference_book,reference_sheet,yellow )
	hilited_columns=get_highlighted_cell_cols_in_sheet(reference_book,reference_sheet,yellow)
	hilited_rows=get_highlighted_cell_rows_in_sheet(reference_book,reference_sheet,yellow)
	nhighlights = how_many_highlighted_cells_in_sheet( reference_book,reference_sheet,yellow )
	stop_string_array = get_highlighted_cell_values_in_sheet(reference_book,reference_sheet,orange)
	sheet_weights = get_highlighted_cell_values_in_sheet( reference_book,reference_sheet,blue )

	total_rubric_pts = 0
	rubric_pts_array = []
	output_array = []
	print "Analyzing rubric for",reference_sheet,"sheet in solution Excel file..."
	for i in range(0,nhighlights):	
		comment = get_comment(reference_book,reference_sheet,letrow_from_colrow_nn(hilited_columns[i],hilited_rows[i]))
		rubric_pts_array.append(float(comment))
		total_rubric_pts = total_rubric_pts + float(comment)
	if is_around(100.0,total_rubric_pts,0.001) == False:
		print "Warning: Sum of comment values in",reference_sheet," sheet is",total_rubric_pts,", the programm wil proceed"
		#user_response = raw_input("Please enter y or n:")
		#if user_response == "n":
		#	quit()
		#elif user_response == "y":
		#	print "Program proceeds...\n"
	output_array.append(rubric_pts_array)
	output_array.append(stop_string_array)
	output_array.append(sheet_weights[0])
	return output_array
	



def grade_nth_sheet_in_folder(submission_folder,reference_book,reference_sheet_number):
	reference_sheet=get_nth_sheet_bn(reference_book,reference_sheet_number)
	list_of_all_students_earned_pts_arrays = []
	
	analysis_array = analyze_reference_sheet(reference_book, reference_sheet)
	rubric_pts_array = analysis_array[0]
	stop_string_array = analysis_array[1]
	sheet_weight = analysis_array[2]

	total_rubric_pts = sum_array(rubric_pts_array)
	names_array = []
	i=0
	for subdir, dirs, files in os.walk(submission_folder):
	    for file in sorted(files):
		ext = os.path.splitext(file)[-1].lower()
		if ext == ".xlsx":
			i = i + 1
			book = file 
			one_individual_first_last_array = get_students_first_last(book)
			appended_first_last = one_individual_first_last_array[0] + " " + one_individual_first_last_array[1]
			names_array.append(appended_first_last)
			sheet=get_nth_sheet_bn(book,reference_sheet_number)
			students_earned_pts_array = grade_sheet_unique_values(book,sheet,reference_book, reference_sheet,rubric_pts_array,total_rubric_pts,stop_string_array)
			students_earned_points_sum = sum_array(students_earned_pts_array)
			print "Student's points: ",students_earned_pts_array,"\n"
			list_of_all_students_earned_pts_arrays.append((appended_first_last,students_earned_points_sum,total_rubric_pts,list(students_earned_pts_array),list(rubric_pts_array),reference_sheet,sheet_weight))
	for i in range(0,len(list_of_all_students_earned_pts_arrays)):
		print i, list_of_all_students_earned_pts_arrays[i][0],list_of_all_students_earned_pts_arrays[i][1],"/",list_of_all_students_earned_pts_arrays[i][2],list_of_all_students_earned_pts_arrays[i][3]
	print
	return list_of_all_students_earned_pts_arrays





def grade_sheet_unique_values(book,sheet,reference_book,reference_sheet,rubric_pts_array,total_rubric_pts,stop_string_array):
	hilites = get_highlighted_cell_values_in_sheet( reference_book,reference_sheet,yellow )
	nhighlights = how_many_highlighted_cells_in_sheet( reference_book,reference_sheet,yellow )
	get_highlighted_cell_letrows_in_sheet( book,sheet,yellow )
	sheets_total_earned_points = total_rubric_pts
	sheets_earned_points_array = []
	for i in range(0,len(hilites)):
		sheets_earned_points_array.append(0.0)
	name = get_students_first_last(book)
	firstName = name[0]
	lastName = name[1]
	print "Grading ",sheet," for : ", firstName, lastName
	for i in range(0,nhighlights):
		if is_around(hilites[i],reserved_for_chart,tolrnc_percent):
				sheet_number = get_sheet_number(book, sheet)
				chart_grade = ChartFinder.grade_chart(reference_book,book,sheet_number)
				sheets_total_earned_points = sheets_total_earned_points - rubric_pts_array[i] * (100.0 - chart_grade) / 100.0
				print "Highlight ",i,": ",rubric_pts_array[i] * (chart_grade) / 100.0,"/",rubric_pts_array[i], " : " ,chart_grade,"/",100.0
				sheets_earned_points_array[i] = rubric_pts_array[i] * (chart_grade) / 100.0
		else:		
			if how_many_match_in_sheet(book,sheet,hilites[i],stop_string_array ) == 0:
				sheets_total_earned_points = sheets_total_earned_points - rubric_pts_array[i]
				print "Highlight ",i,": ","0","/",rubric_pts_array[i]	
				sheets_earned_points_array[i] = 0.0
			else:
				print "Highlight ",i,": ",rubric_pts_array[i],"/",rubric_pts_array[i]
				sheets_earned_points_array[i] = rubric_pts_array[i]
	print "Sheet grade: ",sheets_total_earned_points," / ",total_rubric_pts
	print
	return sheets_earned_points_array



def print_all_grades_per_sheet(master_table):
	print
	num_students = len(master_table[0])
	num_sheets = len(master_table)
	for iStud in range(0,num_students):
		print 
		print iStud,master_table[0][iStud][0],
		for iSheet in range(0,num_sheets):
			print master_table[iSheet][iStud][1],
	print
	print




def print_sheet_dom(master_table):
	print
	num_students = len(master_table[0])
	num_sheets = len(master_table)
	for iSheet in range(0,num_sheets):
		print master_table[iSheet][0][5], " rubric: ", master_table[iSheet][0][4]
		for iStud in range(0,num_students):
			print master_table[iSheet][iStud][0], master_table[iSheet][iStud][1],"/",master_table[iSheet][iStud][2],master_table[iSheet][iStud][3]
		print
	print




def print_student_dom(master_table):
	print
	num_students = len(master_table[0])
	num_sheets = len(master_table)
	for iStud in range(0,num_students):
		student_total_grade = 0.0
		print master_table[0][iStud][0],"(Index: ",iStud,")"
		for iSheet in range(0,num_sheets):
			print master_table[iSheet][iStud][5], master_table[iSheet][iStud][1],"/",master_table[iSheet][iStud][2],"( %",master_table[iSheet][iStud][6],")",master_table[iSheet][iStud][3]
			student_total_grade = student_total_grade + master_table[iSheet][iStud][1]/master_table[iSheet][iStud][2]*master_table[iSheet][iStud][6]
		print ("Total grade: " "%.2f" % student_total_grade)
		print
	print




def grade_all_books_in_folder(submission_folder,reference_book):
	master_table = []
	for iSheet in range(1,count_sheets_b(reference_book)+1):
	#for iSheet in range(5,6):
		master_table.append(grade_nth_sheet_in_folder(submission_folder,reference_book,iSheet))
	for  i in range(1,100):
		print "1: Print total scores"
		print "2: Print list of perormance per sheet"
		print "3: Print list of perormance per student"
		print "q: Exit"
		user_response = raw_input("Choose an option to proceed: \n")
		if user_response == "1":
			print "Printing total scores...\n"
			print_all_grades_per_sheet(master_table)
		elif user_response == "2":
			print_sheet_dom(master_table)
		elif user_response == "3":
			print_student_dom(master_table)
		elif user_response == "q":
			quit()
		else:
			print "Please choose a valid option: \n"

		



	




#solution_file = "ToolsXlsAssignment08Soln1c.xlsx"
#solution_file = "ToolsXlsAssignment05cSoln1b.xlsx"
#solution_file = "ToolsXlsAssignment05cSoln1b-branch1.xlsx"
#solution_file = "ToolsXlsMidterm17.xlsx"
#solution_file = "ToolsXlsMidterm17-P3-alt.xlsx"
#solution_file = "ToolsXlsAssignment11.xlsx"
#solution_file = "ToolsXlsMidterm17-section13.xlsx"
#solution_file = "ToolsXlsAssignment09Soln.xlsx"
#solution_file = "ToolsXlsAssignment11-weightTest.xlsx"
#solution_file = "ToolsXlsAssignment14.xlsx"
#solution_file = "ToolsXlsExam17mon-cheat.xlsx"
#solution_file = "ToolsXlsExam17fri.xlsx"
#solution_file = "ToolsXlsExam17mon-section13.xlsx"
#solution_file = "ToolsXlsExam15.xlsx"
#solution_file = "ToolsXlsExam14-15-16.xlsx"
solution_file = "Solution.xlsx"


############################################################
######This piece of code is used to grade all the books#####
submission_folder = os.getcwd()
reference_book=submission_folder + "/../../Solutions/" + solution_file
grade_all_books_in_folder(submission_folder,reference_book)
############################################################



############################################################
######This piece of code is used to grade nth sheets in all the books#####
#submission_folder = os.getcwd()
#reference_book = reference_book=submission_folder + "/../../Solutions/" + solution_file
#nth = 1
#sheet1_grades=grade_nth_sheet_in_folder(submission_folder,reference_book,nth)
############################################################



############################################################
######This piece of code is used to list all the sheets in all books#####
#submission_folder = os.getcwd()
#list_sheets_all_books(submission_folder )
############################################################


















#wb_data_only_false = load_workbook(filename = "ToolsXlsAssignment11-weightTe.xlsx", data_only=False,read_only=True)
#wb_data_only_true = load_workbook(filename = "1.xlsx", data_only=True,read_only=True)
#ws1 = wb_data_only_false.get_sheet_by_name("Ex4")
#ws2 = wb_data_only_true.get_sheet_by_name("Ex 2")


#print ws2["K15"].value
#if ws1["K15"].value == ws2["K15"].value:
#	print "BAD"
#else:
#	print "GOOD"#


#print get_cell_value("1.xlsx","Ex 2","K15")
#print get_cell_formula("1.xlsx","Ex 2","K15")
#print is_cell_font_colored( "1.xlsx","Ex 2","K15",red )
#print ws1["A44"].font.color.rgb
#print ws1["H1"].fill.start_color.index
#stop_string_array = get_highlighted_cell_values_in_sheet("ToolsXlsMidterm17-section13.xlsx","P1",orange)
#print stop_string_array[0]



#print how_many_none_numeric_match_in_sheet(book,sheet,"=D22*ConvFactor")
#print ws["D11"].value

#print is_cell_highlighted("1.xlsx","Ex 1","B17")
#find_highlighted_cells( "1.xlsx")
#find_highlighted_cells_in_sheet( "1.xlsx","Ex 1")
#find_red_font_in_sheet( "1.xlsx","Ex 1")
#find_red_font( "1.xlsx")
#get_comment( "1.xlsx","Ex 1","E25")
#print get_comment( "1.xlsx","Ex 1","E25")
#print count_sheets_b("1.xlsx")
#print get_nth_sheet_bn("1.xlsx",2)


#a = colrow_from_letrow_s("B17")
#b = letrow_from_colrow_nn(a[0],a[1])
#print is_cell_highlighted("1.xlsx","Ex 1",b)







